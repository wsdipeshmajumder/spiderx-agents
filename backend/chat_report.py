"""Client-ready XLSX performance report for the Chat channel (Conversations tab
→ Export). Two sheets:

  • Summary       — a branded, one-page KPI overview an operator can forward to
                    a client as-is (period, volumes, resolution, satisfaction,
                    leads captured, outcome breakdown).
  • Conversations — one styled row per chat (date, outcome, duration, rating,
                    handoff, captured lead info, summary).

Pure-Python via openpyxl (no system libs). The generator is defensive: any
missing/odd field degrades to a blank cell rather than raising — an export must
never 500 on one weird row.
"""
from __future__ import annotations

import io
import json
import re
from collections import Counter
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Fields inside `extracted` that are meta, not captured lead info.
_META_KEYS = {"csat", "csat_comment", "handoff_requested", "handoff_reason", "_provenance"}


def _prov(ex: dict) -> dict:
    p = ex.get("_provenance")
    return p if isinstance(p, dict) else {}


def _parse_transcript(raw) -> list[dict]:
    """`calls.transcript` is a JSON string of {role, text} turns (or a legacy
    plain string). Returns a clean list of {role, text}."""
    if isinstance(raw, list):
        seq = raw
    elif isinstance(raw, str) and raw.strip():
        try:
            seq = json.loads(raw)
        except Exception:  # noqa: BLE001
            return [{"role": "model", "text": raw.strip()}]
    else:
        return []
    out = []
    if isinstance(seq, list):
        for t in seq:
            if isinstance(t, dict) and t.get("text"):
                out.append({"role": (t.get("role") or "model").strip().lower(),
                            "text": str(t.get("text")).strip()})
    return out

_HEX_RE = re.compile(r"^#?[0-9a-fA-F]{6}$")
_DEFAULT_ACCENT = "4F46E5"   # indigo — matches the widget default


def _accent(agent: dict) -> str:
    """Brand colour (ARGB-friendly 6-hex, no #) from chat_settings.accent_color."""
    cs = agent.get("chat_settings") if isinstance(agent.get("chat_settings"), dict) else {}
    raw = (cs.get("accent_color") or "").strip()
    if _HEX_RE.match(raw):
        return raw.lstrip("#").upper()
    return _DEFAULT_ACCENT


def _display_name(agent: dict) -> str:
    cs = agent.get("chat_settings") if isinstance(agent.get("chat_settings"), dict) else {}
    return (cs.get("display_name") or "").strip() or (agent.get("name") or "Assistant")


def _business(agent: dict) -> str:
    v = agent.get("variables") if isinstance(agent.get("variables"), dict) else {}
    return (v.get("business_name") or "").strip() or (agent.get("name") or "")


def _fmt_dt(dt: Any) -> str:
    if not isinstance(dt, datetime):
        return ""
    return dt.strftime("%d %b %Y, %H:%M")


def _fmt_date(dt: Any) -> str:
    if not isinstance(dt, datetime):
        return ""
    return dt.strftime("%d %b %Y")


def _fmt_dur(sec: Any) -> str:
    try:
        s = int(round(float(sec)))
    except (TypeError, ValueError):
        return ""
    if s <= 0:
        return ""
    m, r = divmod(s, 60)
    return f"{m}m {r:02d}s" if m else f"{r}s"


def _captured(ex: dict) -> dict:
    return {k: v for k, v in ex.items() if k not in _META_KEYS and v not in (None, "", [])}


def _flatten_captured(ex: dict) -> str:
    parts = []
    for k, v in _captured(ex).items():
        label = str(k).replace("_", " ").strip()
        val = ", ".join(map(str, v)) if isinstance(v, list) else (
            str(v) if not isinstance(v, dict) else "; ".join(f"{a}: {b}" for a, b in v.items())
        )
        parts.append(f"{label}: {val}")
    return "  •  ".join(parts)


def _rating(ex: dict) -> str:
    c = ex.get("csat")
    return {"up": "Positive", "down": "Needs work"}.get(c, "")


def build_chat_report_xlsx(
    agent: dict,
    calls: list[dict],
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    generated_at: Optional[datetime] = None,
    include_transcript: bool = False,
) -> bytes:
    """Render the report and return the .xlsx bytes. With `include_transcript`,
    appends a third "Chat log" sheet holding every message of every chat."""
    generated_at = generated_at or datetime.now(timezone.utc)
    accent = _accent(agent)
    business = _business(agent)
    bot = _display_name(agent)

    # ── styles ──────────────────────────────────────────────────────────────
    white = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    title_font = Font(name="Calibri", color="FFFFFF", bold=True, size=18)
    sub_font = Font(name="Calibri", color="FFFFFF", size=11)
    label_font = Font(name="Calibri", color="5B6070", size=10, bold=True)
    metric_font = Font(name="Calibri", color="1A1C25", size=20, bold=True)
    small = Font(name="Calibri", color="6B7080", size=9)
    accent_fill = PatternFill("solid", fgColor=accent)
    card_fill = PatternFill("solid", fgColor="F5F6FA")
    zebra = PatternFill("solid", fgColor="F7F8FC")
    thin = Side(style="thin", color="E6E8EF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    left_mid = Alignment(horizontal="left", vertical="center")

    # ── KPIs ────────────────────────────────────────────────────────────────
    total = len(calls)
    handoffs = 0
    csat_up = csat_down = 0
    leads = 0
    durations = []
    outcomes: Counter = Counter()
    devices: Counter = Counter()
    sources: Counter = Counter()
    period_min = period_max = None
    for c in calls:
        ex = c.get("extracted") if isinstance(c.get("extracted"), dict) else {}
        pv = _prov(ex)
        if pv.get("device"):
            devices[pv["device"]] += 1
        if pv.get("source"):
            sources[pv["source"]] += 1
        if ex.get("handoff_requested") or c.get("outcome") == "transferred_human":
            handoffs += 1
        if ex.get("csat") == "up":
            csat_up += 1
        elif ex.get("csat") == "down":
            csat_down += 1
        if _captured(ex):
            leads += 1
        d = c.get("duration_s")
        if isinstance(d, (int, float)) and d > 0:
            durations.append(d)
        outcomes[(c.get("outcome") or "unknown")] += 1
        st = c.get("started_at")
        if isinstance(st, datetime):
            period_min = st if period_min is None or st < period_min else period_min
            period_max = st if period_max is None or st > period_max else period_max

    rated = csat_up + csat_down
    resolved_ai = total - handoffs
    avg_dur = (sum(durations) / len(durations)) if durations else 0
    period_lo = date_from or period_min
    period_hi = date_to or period_max
    period_str = (
        f"{_fmt_date(period_lo)} – {_fmt_date(period_hi)}"
        if period_lo and period_hi else "All time"
    )

    wb = Workbook()

    # ══ Sheet 1 · Summary ═════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    widths = [3, 24, 22, 22, 22, 3]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Banner (rows 1-3, cols B..E)
    ws.merge_cells("B1:E1")
    ws["B1"] = f"{business or bot} — Chat Performance Report"
    ws["B1"].font = title_font
    ws["B1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.merge_cells("B2:E2")
    ws["B2"] = f"AI assistant: {bot}    ·    Period: {period_str}"
    ws["B2"].font = sub_font
    ws.merge_cells("B3:E3")
    ws["B3"] = f"Generated {generated_at.strftime('%d %b %Y')}    ·    Powered by SpiderX.AI"
    ws["B3"].font = sub_font
    for r in (1, 2, 3):
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = accent_fill
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # KPI cards (2 rows × 3). (label, value)
    sat = f"{round(100 * csat_up / rated)}%" if rated else "—"
    cards = [
        ("Total conversations", str(total)),
        ("Resolved by AI", str(resolved_ai)),
        ("Human handoffs", str(handoffs)),
        ("Positive rating", sat),
        ("Leads / info captured", str(leads)),
        ("Avg conversation length", _fmt_dur(avg_dur) or "—"),
    ]
    start_row = 5
    positions = [("B", "C"), ("C", "C"), ("D", "D")]  # placeholder; compute below
    col_pairs = [("B",), ("C",), ("D",), ("E",)]
    # Lay 3 cards per row across B,C,D (E is spacer-narrow) — use B, C, D, and
    # wrap; simplest: map card index → (row, col) over cols B,C,D.
    grid_cols = ["B", "C", "D"]
    for idx, (label, value) in enumerate(cards):
        gr = start_row + (idx // 3) * 4
        gc = grid_cols[idx % 3]
        lab = ws[f"{gc}{gr}"]
        lab.value = label.upper()
        lab.font = label_font
        lab.fill = card_fill
        lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        val = ws[f"{gc}{gr + 1}"]
        val.value = value
        val.font = metric_font
        val.fill = card_fill
        val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[gr].height = 20
        ws.row_dimensions[gr + 1].height = 30
        for rr in (gr, gr + 1):
            ws[f"{gc}{rr}"].border = border

    # Breakdown tables (outcome, then visitor device + top sources for
    # analytics). Each: bold title, accent header row, zebra-striped data.
    def _table(row: int, title: str, first_col: str, pairs: list) -> int:
        ws[f"B{row}"] = title
        ws[f"B{row}"].font = Font(name="Calibri", bold=True, size=12, color="1A1C25")
        h = row + 1
        for col, name in zip(("B", "C", "D"), (first_col, "Count", "Share")):
            cell = ws[f"{col}{h}"]
            cell.value = name
            cell.font = white
            cell.fill = accent_fill
            cell.alignment = left_mid if col == "B" else center
            cell.border = border
        r = h + 1
        if not pairs:
            ws[f"B{r}"] = "No data in this period."
            ws[f"B{r}"].font = small
            for col in ("B", "C", "D"):
                ws[f"{col}{r}"].border = border
            return r + 2
        for label, n in pairs:
            ws[f"B{r}"] = str(label).replace("_", " ")
            ws[f"C{r}"] = n
            ws[f"D{r}"] = (n / total) if total else 0
            ws[f"D{r}"].number_format = "0%"
            ws[f"B{r}"].alignment = left_mid
            ws[f"C{r}"].alignment = center
            ws[f"D{r}"].alignment = center
            for col in ("B", "C", "D"):
                ws[f"{col}{r}"].border = border
                if (r - h) % 2 == 0:
                    ws[f"{col}{r}"].fill = zebra
            r += 1
        return r + 1   # blank spacer row after the table

    ri = _table(start_row + 8, "Outcome breakdown", "Outcome", outcomes.most_common())
    ri = _table(ri, "Visitor device", "Device", devices.most_common())
    ri = _table(ri, "Top sources", "Source", sources.most_common(8))

    note = ws.cell(row=ri + 1, column=2)
    note.value = ("“Resolved by AI” = conversations handled end-to-end without a human handoff.  "
                  "“Positive rating” = share of rated chats the visitor marked 👍.")
    note.font = small

    # ══ Sheet 2 · Conversations ═══════════════════════════════════════════════
    ws2 = wb.create_sheet("Conversations")
    ws2.sheet_view.showGridLines = False
    headers = ["Date", "Time", "Outcome", "Duration", "Rating", "Handoff", "Device", "Source", "Captured info", "Summary"]
    cwidths = [14, 8, 20, 11, 12, 10, 12, 22, 40, 56]
    for i, w in enumerate(cwidths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for i, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = white
        cell.fill = accent_fill
        cell.alignment = center if h in ("Duration", "Rating", "Handoff", "Device") else left_mid
        cell.border = border
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"

    r = 2
    for c in calls:
        ex = c.get("extracted") if isinstance(c.get("extracted"), dict) else {}
        pv = _prov(ex)
        st = c.get("started_at")
        row_vals = [
            _fmt_date(st),
            st.strftime("%H:%M") if isinstance(st, datetime) else "",
            str(c.get("outcome") or "unknown").replace("_", " "),
            _fmt_dur(c.get("duration_s")),
            _rating(ex),
            "Yes" if (ex.get("handoff_requested") or c.get("outcome") == "transferred_human") else "",
            pv.get("device", ""),
            pv.get("source", ""),
            _flatten_captured(ex),
            (c.get("summary") or "").strip(),
        ]
        for i, v in enumerate(row_vals, start=1):
            cell = ws2.cell(row=r, column=i, value=v)
            cell.border = border
            cell.alignment = center if i in (4, 5, 6, 7) else wrap_top
            if r % 2 == 0:
                cell.fill = zebra
        r += 1
    if not calls:
        ws2.cell(row=2, column=1, value="No conversations in this period.").font = Font(italic=True, color="6B7080")

    # ══ Sheet 3 · Full chat log (opt-in) ═════════════════════════════════════
    # Every message of every conversation, grouped chat-by-chat with a tinted
    # header row (date · outcome · captured info) so it reads as a transcript.
    if not include_transcript:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    ws3 = wb.create_sheet("Chat log")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 96
    speaker_font = Font(name="Calibri", bold=True, size=10, color="5B6070")
    visitor_font = Font(name="Calibri", bold=True, size=10, color=accent)
    msg_font = Font(name="Calibri", size=10.5, color="1A1C25")
    grp_fill = PatternFill("solid", fgColor=accent)
    grp_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    r = 1
    any_tx = False
    for idx, c in enumerate(calls, start=1):
        turns = _parse_transcript(c.get("transcript"))
        if not turns:
            continue
        any_tx = True
        st = c.get("started_at")
        head = f"Chat {idx}  ·  {_fmt_dt(st)}  ·  {str(c.get('outcome') or 'unknown').replace('_', ' ')}"
        cap = _flatten_captured(c.get("extracted") if isinstance(c.get("extracted"), dict) else {})
        if cap:
            head += f"  ·  {cap}"
        ws3.merge_cells(f"A{r}:B{r}")
        gc = ws3.cell(row=r, column=1, value=head)
        gc.fill = grp_fill
        gc.font = grp_font
        gc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws3.row_dimensions[r].height = 22
        r += 1
        for t in turns:
            is_user = t["role"] in ("user", "caller", "human")
            sp = ws3.cell(row=r, column=1, value="Visitor" if is_user else bot)
            sp.font = visitor_font if is_user else speaker_font
            sp.alignment = Alignment(vertical="top", horizontal="left")
            mc = ws3.cell(row=r, column=2, value=t["text"])
            mc.font = msg_font
            mc.alignment = wrap_top
            if not is_user:
                for col in (1, 2):
                    ws3.cell(row=r, column=col).fill = zebra
            r += 1
        r += 1   # blank row between chats
    if not any_tx:
        ws3.cell(row=1, column=1, value="No transcripts in this period.").font = Font(italic=True, color="6B7080")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
